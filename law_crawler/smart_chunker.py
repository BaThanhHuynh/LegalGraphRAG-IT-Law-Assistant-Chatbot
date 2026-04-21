"""
smart_chunker.py
================
Xử lý chunking thông minh cho dữ liệu luật pháp tiếng Việt.

Chiến lược 4 tầng:
  Tầng 1: Điều ≤ 450 ký tự  → dùng thẳng
  Tầng 2: Điều > 450 ký tự  → split theo Khoản (1. 2. 3.)
  Tầng 3: Khoản > 450 ký tự → split theo Điểm (a) b) c))
  Tầng 4: vẫn còn dài        → RecursiveCharacterTextSplitter / simple_split

Xử lý đặc biệt:
  - Điều bị bãi bỏ trong VBHN (toàn bộ = "(Bãi bỏ)") → Tier 0
  - Nội dung bị truncate Excel → đọc lại từ DOCX gốc (xem patch_truncated_records)
  - Sau split: lọc bỏ chunk rác (< 5 ký tự hoặc là số đơn thuần)
  - Tiêu đề điều tách rời (do RE_KHOAN) → merge vào khoản 1



Cách dùng:
    python smart_chunker.py --input data/law_data_output.xlsx --output data/law_chunks --format both

    # Tuỳ chỉnh chunk size
    python smart_chunker.py --input data/law_data_output.xlsx --output data/law_chunks --chunk_size 400 --overlap 40

Yêu cầu:
    pip install openpyxl
    pip install langchain-text-splitters   # tuỳ chọn, có thì tốt hơn
"""

import re
import sys
import json
import hashlib
import logging
import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Thiếu thư viện: pip install openpyxl")
    sys.exit(1)

# LangChain tuỳ chọn — chỉ dùng langchain_text_splitters (có 's'), KHÔNG dùng
# langchain.text_splitter (cũ, deprecated từ v0.2+)
try:
    from langchain_text_splitters import RecursiveCharacterTextSplitter
    LANGCHAIN_OK = True
except ImportError:
    LANGCHAIN_OK = False

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger(__name__)

# ── Cấu hình mặc định ────────────────────────────────────────────────────────
CHUNK_SIZE    = 400    # ký tự tối đa mỗi chunk (PhoBERT max ~512 ký tự VN)
CHUNK_OVERLAP = 50     # ký tự overlap giữa chunk liền kề (Tier 4)
EXCEL_MAX_LEN = 32767  # giới hạn ký tự của 1 cell Excel
MIN_USEFUL_LEN = 20     # chunk ngắn hơn mức này là rác, sẽ bị lọc bỏ

# Regex Điều — dùng để mask trước khi split khoản
_RE_DIEU_DOT = re.compile(r"(Điều\s+\d+)\.", re.IGNORECASE)

# Regex Khoản: "1. ", "2. "... — phải đứng sau mask để tránh khớp "Điều 2."
_RE_KHOAN = re.compile(r"(?<!\w)(\d{1,2})\.\s+(?=[^\d\s])")

# Regex Điểm: "a) ", "b) ", "đ) "...
_RE_DIEM = re.compile(r"(?<![a-zA-ZÀ-ỹ])([a-zđ])\)\s+")

# Regex số đơn thuần để lọc chunk rác
_RE_ONLY_NUMBER = re.compile(r"^\d{1,3}$")

# Regex điều bị bãi bỏ trong VBHN: toàn bộ nội dung = "(Bãi bỏ)" hoặc "(Hết hiệu lực)"
_RE_WHOLE_REPEALED = re.compile(
    r"^\s*\((bãi bỏ|hết hiệu lực|được bãi bỏ|bị bãi bỏ)\)\s*$",
    re.IGNORECASE
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def make_chunk_id(base_id: str, sub_index: int) -> str:
    raw = f"{base_id}_sub{sub_index}"
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def is_repealed(text: str) -> bool:
    """
    True khi TOÀN BỘ nội dung điều = annotation bãi bỏ VBHN.

    Ví dụ True:  "(Bãi bỏ)"  |  "(Hết hiệu lực)"  |  "" (trống)
    Ví dụ False: "...yêu cầu xóa bỏ thông tin..."  (xóa bỏ trong nghiệp vụ)
                 "...điểm a (được bãi bỏ); b) Tịch thu..."  (điểm con bị bỏ)
    """
    stripped = text.strip()
    # Trống hoặc quá ngắn (< 5 ký tự) → chắc chắn không có nội dung
    if len(stripped) < 5:
        return True
    # Toàn bộ = annotation VBHN chuẩn
    if _RE_WHOLE_REPEALED.match(stripped):
        return True
    return False


_RE_KHOAN_REPEALED = re.compile(
    r"^(\d{1,3})\.\s*\((bãi bỏ|hết hiệu lực|được bãi bỏ)\)\s*$",
    re.IGNORECASE
)


def is_junk_chunk(text: str) -> bool:
    stripped = text.strip()
    if len(stripped) < MIN_USEFUL_LEN:
        return True
    if _RE_ONLY_NUMBER.match(stripped):
        return True
    # Thêm lọc chunk rác kiểu "1.", "2.", "a)", "đ)" 
    if re.match(r"^[0-9]+\.$", stripped) or re.match(r"^[a-zđ]\)\s*$", stripped, re.IGNORECASE):
        return True
    return False


def is_khoan_repealed(text: str) -> bool:
    """True nếu đây là khoản bị bãi bỏ: 'X. (được bãi bỏ)'."""
    return bool(_RE_KHOAN_REPEALED.match(text.strip()))


def build_prefix(row: dict) -> str:
    """
    Tạo prefix ngữ cảnh cho mỗi chunk, dùng làm đầu vào embedding.
    Format: "[Tên VB (Số hiệu)] | [Chương X: Tên chương] | [Điều Y. Tên điều]"
    """
    parts = []
    ten_vb  = row.get("ten_van_ban", "").strip()
    so_hieu = row.get("so_hieu", "").strip()
    if ten_vb:
        parts.append(f"{ten_vb} ({so_hieu})" if so_hieu else ten_vb)

    chuong_so  = row.get("chuong_so", "").strip()
    chuong_ten = row.get("chuong_ten", "").strip()
    if chuong_ten:
        parts.append(f"Chương {chuong_so}: {chuong_ten}" if chuong_so else chuong_ten)

    dieu_so  = row.get("dieu_so", "").strip()
    dieu_ten = row.get("dieu_ten", "").strip()
    if dieu_so:
        parts.append(f"Điều {dieu_so}. {dieu_ten}" if dieu_ten else f"Điều {dieu_so}")

    return " | ".join(parts)


# ── Split functions ───────────────────────────────────────────────────────────

def split_by_khoan(text: str) -> list:
    """
    Split nội dung điều theo Khoản (1. 2. 3. ...).

    FIX BUG: RE_KHOAN cũ match cả "Điều 2." → tách "Điều" thành chunk 4 ký tự.
    Giải pháp: mask "Điều X." → "Điều X[DOT]" trước khi tìm khoản,
               rồi restore sau khi split xong.

    Sau khi split: nếu chunk đầu tiên không bắt đầu bằng số (là tiêu đề điều)
    → merge vào chunk khoản 1 để giữ ngữ cảnh.
    """
    # Bước 1: Mask để tránh nhầm số trong "Điều X." là khoản
    safe = _RE_DIEU_DOT.sub(r"\1[DOT]", text)
    # Cũng mask Mục X., Chương X. cho an toàn
    safe = re.sub(r"((?:Mục|Chương|Khoản|Tiết)\s+[\dIVXLCDM]+)\.", r"\1[DOT]", safe, flags=re.IGNORECASE)

    positions = [m.start() for m in _RE_KHOAN.finditer(safe)]
    if not positions:
        return [text]  # không có khoản → trả về nguyên

    # Bước 2: Split tại vị trí khoản
    raw = []
    prev = 0
    for pos in positions:
        if pos > prev:
            chunk = safe[prev:pos].strip()
            if chunk:
                raw.append(chunk.replace("[DOT]", "."))
        prev = pos
    if prev < len(safe):
        chunk = safe[prev:].strip()
        if chunk:
            raw.append(chunk.replace("[DOT]", "."))

    if not raw:
        return [text]

    # Bước 3: Merge tiêu đề điều (chunk đầu không bắt đầu bằng số) vào khoản 1
    _RE_STARTS_NUM = re.compile(r"^\d{1,2}\.")
    if len(raw) > 1 and not _RE_STARTS_NUM.match(raw[0]):
        merged_first = f"{raw[0]} {raw[1]}"
        return [merged_first] + raw[2:]

    return raw


def split_by_diem(text: str) -> list:
    """Split khoản theo Điểm (a) b) c) đ) ...)."""
    positions = [m.start() for m in _RE_DIEM.finditer(text)]
    if not positions:
        return [text]

    chunks = []
    prev = 0
    for pos in positions:
        if pos > prev:
            chunk = text[prev:pos].strip()
            if chunk:
                chunks.append(chunk)
        prev = pos
    if prev < len(text):
        chunk = text[prev:].strip()
        if chunk:
            chunks.append(chunk)

    return chunks if chunks else [text]


def greedy_merge_diem(diem_list: list, chunk_size: int) -> list:
    """
    Gộp các điểm a,b,c,d liên tiếp lại thành nhóm có độ dài ≤ chunk_size.

    Lý do: split_by_diem() tách từng điểm riêng lẻ → nhiều chunk ngắn mất ngữ cảnh.
    Ví dụ: ['a) Thẩm định;', 'b) Đánh giá;', 'c) Kiểm tra;']
    → Gộp: 'a) Thẩm định; b) Đánh giá; c) Kiểm tra;' (1 chunk)

    Điểm đầu tiên thường là tiêu đề/khoản chứa dấu "bao gồm:" → giữ riêng
    nếu đủ dài, gộp các điểm tiếp theo vào.
    """
    if not diem_list:
        return []
    if len(diem_list) == 1:
        return diem_list

    result = []
    current = diem_list[0]

    for piece in diem_list[1:]:
        candidate = current + " " + piece
        if len(candidate) <= chunk_size:
            current = candidate  # gộp vào
        else:
            result.append(current)
            current = piece

    result.append(current)
    return result


def simple_split(text: str, chunk_size: int, overlap: int) -> list:
    """
    Fallback khi không có LangChain.
    Split tại dấu câu gần nhất trước chunk_size.

    Fix: điểm cắt phải ở ít nhất nửa chunk (tránh tạo mảnh vỡ overlap nhỏ).
    Overlap chỉ áp dụng khi chunk đủ dài (≥ 100 ký tự).
    """
    if len(text) <= chunk_size:
        return [text]

    separators = ["\n\n", "\n", "; ", ". ", ", ", " "]
    chunks = []
    start = 0
    min_cut_pos = chunk_size // 2  # điểm cắt phải ở ít nhất 1/2 chunk

    while start < len(text):
        end = min(start + chunk_size, len(text))
        if end == len(text):
            tail = text[start:].strip()
            if tail:
                chunks.append(tail)
            break
        cut = end
        for sep in separators:
            pos = text.rfind(sep, start + min_cut_pos, end)
            if pos > start:
                cut = pos + len(sep)
                break
        piece = text[start:cut].strip()
        if piece:
            chunks.append(piece)
        # Overlap chỉ khi piece đủ dài, tránh tạo mảnh vỡ
        effective_overlap = overlap if len(piece) >= 100 else 0
        start = max(cut - effective_overlap, cut if effective_overlap == 0 else start + 1)
        if start >= len(text):
            break

    return [c for c in chunks if c.strip()]


def tier4_split(text: str) -> list:
    """Tier 4: LangChain nếu có, fallback simple_split."""
    if LANGCHAIN_OK:
        splitter = RecursiveCharacterTextSplitter(
            chunk_size=CHUNK_SIZE,
            chunk_overlap=CHUNK_OVERLAP,
            separators=["\n\n", "\n", "; ", ". ", ", ", " ", ""],
            length_function=len,
        )
        return splitter.split_text(text)
    return simple_split(text, CHUNK_SIZE, CHUNK_OVERLAP)


# ── Core chunking ─────────────────────────────────────────────────────────────

def chunk_record(row: dict) -> list:
    """
    Xử lý 1 record (1 điều luật) → list các chunk con.

    Returns:
        List[dict] mỗi dict là 1 chunk với metadata đầy đủ.
    """
    noi_dung = str(row.get("noi_dung_dieu", "")).strip()
    original_chunk_id = row.get("chunk_id", "")
    prefix = build_prefix(row)
    is_truncated = row.get("is_truncated_excel", False)

    if is_junk_chunk(noi_dung):
        return []

    # ── Điều bị bãi bỏ ───────────────────────────────────────────────────
    if is_repealed(noi_dung):
        return [_make_chunk(
            row=row,
            text=noi_dung or "(Điều đã được bãi bỏ hoặc hết hiệu lực)",
            chunk_id=original_chunk_id,
            sub_index=0, total_sub=1,
            is_repealed_flag=True, tier=0,
            is_truncated=is_truncated,
        )]

    # ── Cảnh báo Excel truncate ───────────────────────────────────────────
    if len(noi_dung) >= EXCEL_MAX_LEN:
        log.warning(
            f"  ⚠ Chunk có thể bị truncate (Excel limit): "
            f"{row.get('source_file','')} Điều {row.get('dieu_so','')}"
        )

    # ── Tier 1: ngắn, dùng thẳng ─────────────────────────────────────────
    if len(noi_dung) <= CHUNK_SIZE:
        ctx = f"{prefix} | {noi_dung}" if prefix else noi_dung
        return [_make_chunk(row, noi_dung, original_chunk_id, 0, 1, False, tier=1,
                    context_text=ctx, is_truncated=is_truncated)]

    # ── Tier 2: split theo khoản ──────────────────────────────────────────
    khoan_list = split_by_khoan(noi_dung)
    final_texts = []  # list of (text, tier)

    for khoan in khoan_list:
        khoan = khoan.strip()
        if not khoan:
            continue
        if len(khoan) <= CHUNK_SIZE:
            final_texts.append((khoan, 2))
        else:
            # ── Tier 3: split theo điểm ──────────────────────────────────
            diem_list = split_by_diem(khoan)
            if len(diem_list) > 1:
                # Gộp các điểm a,b,c,d liên tiếp lại cho đến chunk_size
                # thay vì để từng điểm riêng lẻ (tránh chunk ngắn mất ngữ cảnh)
                merged_diem = greedy_merge_diem(diem_list, CHUNK_SIZE)
                for dc in merged_diem:
                    dc = dc.strip()
                    if not dc:
                        continue
                    if len(dc) <= CHUNK_SIZE:
                        final_texts.append((dc, 3))
                    else:
                        # ── Tier 4: recursive ─────────────────────────────
                        for sub in tier4_split(dc):
                            if sub.strip():
                                final_texts.append((sub.strip(), 4))
            else:
                for sub in tier4_split(khoan):
                    if sub.strip():
                        final_texts.append((sub.strip(), 4))

    # ── Lọc chunk rác & đánh flag khoản bị bãi bỏ ───────────────────────
    # Bỏ chunk quá ngắn (< 5 ký tự) hoặc chỉ là số đơn thuần "1", "2", "3"
    # Khoản "X. (được bãi bỏ)" → đánh flag thay vì bỏ đi
    repealed_khoan = []
    clean_texts = []
    for t, tier in final_texts:
        if is_junk_chunk(t):
            continue  # bỏ rác
        if is_khoan_repealed(t):
            repealed_khoan.append(t)  # giữ lại nhưng đánh flag
        else:
            clean_texts.append((t, tier))
    final_texts = clean_texts

    if not final_texts:
        # Fallback: giữ nguyên nội dung gốc nếu mọi chunk đều bị lọc
        return [_make_chunk(row, noi_dung, original_chunk_id, 0, 1, False, tier=1, is_truncated=is_truncated)]

    # ── Tạo chunks với context prefix ────────────────────────────────────
    result = []
    total = len(final_texts)
    for i, (text, tier) in enumerate(final_texts):
        context = f"{prefix} | {text}" if prefix else text
        result.append(_make_chunk(
            row=row,
            text=text,
            chunk_id=make_chunk_id(original_chunk_id, i),
            sub_index=i,
            total_sub=total,
            is_repealed_flag=False,
            tier=tier,
            context_text=context,
            is_truncated=is_truncated,
        ))

    return result


def _make_chunk(row: dict, text: str, chunk_id: str,
                sub_index: int, total_sub: int,
                is_repealed_flag: bool, tier: int,
                context_text: str = None,
                is_truncated: bool = False) -> dict:
    """Tạo dict chunk đầy đủ metadata."""
    chunk = {k: v for k, v in row.items()
             if k not in ("noi_dung_dieu", "do_dai_ky_tu", "chunk_id", "is_truncated_excel")}
    chunk.update({
        "chunk_id":        chunk_id,
        "chunk_sub_index": sub_index,
        "chunk_total_sub": total_sub,
        "chunk_tier":      tier,
        "noi_dung_chunk":  text,
        "do_dai_chunk":    len(text),
        "context_text":    context_text or text,
        "is_repealed":     is_repealed_flag,
        "is_truncated":    is_truncated,
    })
    return chunk


# ── Đọc Excel ─────────────────────────────────────────────────────────────────

HEADER_MAP = {
    "Source File": "source_file", "Tên văn bản": "ten_van_ban",
    "Số hiệu": "so_hieu", "Số VBHN": "so_vbhn",
    "Loại văn bản": "loai_van_ban", "Cơ quan ban hành": "co_quan_ban_hanh",
    "Ngày ban hành": "ngay_ban_hanh", "Ngày hiệu lực": "ngay_hieu_luc",
    "Ngày hết hiệu lực": "ngay_het_hieu_luc", "Trạng thái": "trang_thai",
    "Sửa đổi bởi": "sua_doi_boi", "Bản sử dụng": "ban_su_dung",
    "Nhóm": "nhom", "Ghi chú": "ghi_chu",
    "Chương số": "chuong_so", "Chương tên": "chuong_ten",
    "Mục số": "muc_so", "Mục tên": "muc_ten",
    "Điều số": "dieu_so", "Điều tên": "dieu_ten",
    "Nội dung điều": "noi_dung_dieu",
    "Độ dài (ký tự)": "do_dai_ky_tu", "Chunk ID": "chunk_id", "STT": "stt",
}


def read_excel(path: str) -> list:
    """Đọc sheet 'Dữ liệu luật' từ file Excel."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Dữ liệu luật" not in wb.sheetnames:
        log.error("Không tìm thấy sheet 'Dữ liệu luật'")
        sys.exit(1)

    ws = wb["Dữ liệu luật"]
    rows = list(ws.values)
    if not rows:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    records = []
    for row in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rec = {}
        for i, h in enumerate(headers):
            key = HEADER_MAP.get(h, h.lower().replace(" ", "_"))
            val = row[i] if i < len(row) else None
            rec[key] = str(val).strip() if val is not None else ""
        # Flag truncated
        nd = rec.get("noi_dung_dieu", "")
        rec["is_truncated_excel"] = len(nd) >= EXCEL_MAX_LEN
        records.append(rec)

    log.info(f"  📖 Đọc {len(records):,} điều từ Excel")
    return records


# ── Patch điều bị truncate từ DOCX gốc ───────────────────────────────────────

def patch_truncated_records(records: list, docx_folder: str = None) -> list:
    """
    Với các record bị flag is_truncated_excel=True:
    Đọc lại nội dung đầy đủ từ file DOCX gốc trong docx_folder.

    Cũng xử lý Điều 124 của NĐ 15/2020: loại bỏ footnote bị parse nhầm.

    Args:
        records: list record đọc từ Excel
        docx_folder: đường dẫn folder chứa file DOCX gốc (data/raw/)
    """
    truncated = [r for r in records if r.get("is_truncated_excel")]
    if not truncated:
        return records

    log.info(f"  🔧 Patch {len(truncated)} điều bị truncate từ DOCX gốc...")

    if not docx_folder:
        log.warning("  ⚠ Chưa chỉ định --docx_folder. Bỏ qua patch, dữ liệu có thể bị cắt ngắn.")
        return records

    folder = Path(docx_folder)
    if not folder.exists():
        log.warning(f"  ⚠ Không tìm thấy folder: {folder}")
        return records

    try:
        from docx import Document as DocxDocument
    except ImportError:
        log.warning("  ⚠ Thiếu python-docx. Bỏ qua patch: pip install python-docx")
        return records

    # Cache: tránh đọc cùng 1 DOCX nhiều lần
    docx_cache = {}
    RE_DIEU = re.compile(r"^(điều\s+(\d+)[\.:]?\s*(.*))", re.IGNORECASE)
    RE_FOOTNOTE = re.compile(r"^\d+\s+[A-ZĐ]")  # footnote: "133 Điều này được..."

    def load_docx_dieu(source_file: str) -> dict:
        """Trả về dict {dieu_so: noi_dung} từ DOCX."""
        if source_file in docx_cache:
            return docx_cache[source_file]

        # Tìm file trong folder
        candidates = list(folder.glob(f"**/{source_file}"))
        if not candidates:
            # Thử tìm fuzzy (bỏ qua extension)
            stem = Path(source_file).stem
            candidates = list(folder.glob(f"**/*{stem[:20]}*.docx"))

        if not candidates:
            log.warning(f"  ⚠ Không tìm thấy DOCX: {source_file}")
            docx_cache[source_file] = {}
            return {}

        try:
            doc = DocxDocument(str(candidates[0]))
        except Exception as e:
            log.warning(f"  ⚠ Lỗi đọc {candidates[0]}: {e}")
            docx_cache[source_file] = {}
            return {}

        dieu_dict = {}
        current_dieu = None
        current_paras = []
        stop_footnote = False

        for para in doc.paragraphs:
            text = " ".join(para.text.split()).strip()
            if not text:
                continue
            m = RE_DIEU.match(text)
            if m:
                # Lưu điều trước
                if current_dieu and current_paras:
                    dieu_dict[current_dieu] = " ".join(current_paras)
                current_dieu = m.group(2)
                current_paras = [text]
                stop_footnote = False
            else:
                if current_dieu and not stop_footnote:
                    # Dừng nếu gặp footnote (dòng bắt đầu bằng số + chữ hoa)
                    if RE_FOOTNOTE.match(text):
                        stop_footnote = True
                    else:
                        current_paras.append(text)

        if current_dieu and current_paras:
            dieu_dict[current_dieu] = " ".join(current_paras)

        docx_cache[source_file] = dieu_dict
        log.info(f"    Đọc DOCX {source_file}: {len(dieu_dict)} điều")
        return dieu_dict

    # Patch từng record
    patched = 0
    for rec in records:
        if not rec.get("is_truncated_excel"):
            continue
        source_file = rec.get("source_file", "")
        dieu_so = rec.get("dieu_so", "").strip()
        dieu_dict = load_docx_dieu(source_file)
        if dieu_so in dieu_dict:
            full_content = dieu_dict[dieu_so]
            old_len = len(rec.get("noi_dung_dieu", ""))
            rec["noi_dung_dieu"] = full_content
            rec["is_truncated_excel"] = False
            patched += 1
            log.info(f"    ✅ Patch Điều {dieu_so} ({source_file}): {old_len} → {len(full_content)} ký tự")

    log.info(f"  🔧 Đã patch {patched}/{len(truncated)} điều")
    return records


# ── Xuất Excel ────────────────────────────────────────────────────────────────

def export_excel_chunks(chunks: list, output_path: str):
    """Xuất chunks ra Excel với màu sắc theo tier và 3 sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chunks"
    ws.freeze_panes = "A2"

    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    COLS = [
        ("chunk_id",         "Chunk ID",                  14),
        ("chunk_sub_index",  "Sub Index",                  9),
        ("chunk_total_sub",  "Total Sub",                  9),
        ("chunk_tier",       "Tier",                       6),
        ("is_repealed",      "Bị bãi bỏ?",               10),
        ("source_file",      "Source File",               30),
        ("ten_van_ban",      "Tên văn bản",               35),
        ("so_hieu",          "Số hiệu",                   16),
        ("trang_thai",       "Trạng thái",                16),
        ("ngay_hieu_luc",    "Ngày hiệu lực",             14),
        ("ngay_het_hieu_luc","Ngày hết hiệu lực",         16),
        ("chuong_so",        "Chương số",                 10),
        ("chuong_ten",       "Chương tên",                30),
        ("muc_so",           "Mục số",                     8),
        ("muc_ten",          "Mục tên",                   25),
        ("dieu_so",          "Điều số",                    8),
        ("dieu_ten",         "Điều tên",                  35),
        ("noi_dung_chunk",   "Nội dung chunk",            80),
        ("context_text",     "Context text (→ embedding)",80),
        ("do_dai_chunk",     "Độ dài chunk",              12),
        ("is_truncated",     "Bị cắt ngắn?",             12),
    ]

    TIER_COLORS = {
        0: "F7C6C7",  # đỏ nhạt  = bãi bỏ
        1: "D9EAD3",  # xanh lá  = dùng thẳng
        2: "CFE2F3",  # xanh lam = split khoản
        3: "FFF2CC",  # vàng     = split điểm
        4: "F4CCFF",  # tím      = recursive
    }

    for ci, (_, header, width) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="2E4057")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28

    for ri, chunk in enumerate(chunks, 2):
        tier = int(chunk.get("chunk_tier", 1))
        fill = PatternFill("solid", fgColor=TIER_COLORS.get(tier, "FFFFFF"))
        for ci, (key, _, _) in enumerate(COLS, 1):
            val = chunk.get(key, "")
            if isinstance(val, bool):
                val = "✓" if val else ""
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.border = border
            c.alignment = Alignment(
                vertical="top",
                wrap_text=(key in ("noi_dung_chunk", "context_text"))
            )
            if key == "trang_thai":
                if val == "con_hieu_luc":
                    c.font = Font(color="1A7A44", bold=True)
                elif val == "het_hieu_luc":
                    c.font = Font(color="C0392B", bold=True)
            if key == "is_repealed" and val == "✓":
                c.font = Font(color="C0392B", bold=True)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # Sheet legend
    ws2 = wb.create_sheet("Legend")
    ws2["A1"] = "GIẢI THÍCH MÀU SẮC"
    ws2["A1"].font = Font(bold=True, size=12)
    for i, (label, desc, color) in enumerate([
        ("Tier 0 – Đỏ",   "Điều bị bãi bỏ/hết hiệu lực (toàn bộ = annotation VBHN)", "F7C6C7"),
        ("Tier 1 – Xanh lá","Điều ngắn ≤ 450 ký tự, dùng thẳng",                     "D9EAD3"),
        ("Tier 2 – Xanh lam","Split theo Khoản (1. 2. 3.)",                           "CFE2F3"),
        ("Tier 3 – Vàng", "Split theo Điểm (a) b) c))",                               "FFF2CC"),
        ("Tier 4 – Tím",  "Split Recursive (LangChain hoặc simple_split)",            "F4CCFF"),
    ], start=3):
        ws2.cell(row=i, column=1, value=label).fill = PatternFill("solid", fgColor=color)
        ws2.cell(row=i, column=2, value=desc)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 60

    # Sheet thống kê
    ws3 = wb.create_sheet("Thống kê chunks")
    total = len(chunks)
    by_tier = {}
    for c in chunks:
        t = int(c.get("chunk_tier", 1))
        by_tier[t] = by_tier.get(t, 0) + 1
    lengths = [int(c.get("do_dai_chunk", 0)) for c in chunks if c.get("do_dai_chunk")]
    repealed = sum(1 for c in chunks if c.get("is_repealed"))
    truncated = sum(1 for c in chunks if c.get("is_truncated"))

    ws3["A1"] = "THỐNG KÊ CHUNKS"
    ws3["A1"].font = Font(bold=True, size=12)
    stats = [
        ("Tổng số chunk",           total),
        ("  Tier 1 (nguyên ≤450)",  by_tier.get(1, 0)),
        ("  Tier 2 (split khoản)",  by_tier.get(2, 0)),
        ("  Tier 3 (split điểm)",   by_tier.get(3, 0)),
        ("  Tier 4 (recursive)",    by_tier.get(4, 0)),
        ("  Tier 0 (bãi bỏ)",       by_tier.get(0, 0)),
        ("", ""),
        ("Điều bị bãi bỏ",          repealed),
        ("Chunk từ điều bị truncate",truncated),
        ("", ""),
        ("Min độ dài chunk",         min(lengths) if lengths else 0),
        ("Max độ dài chunk",         max(lengths) if lengths else 0),
        ("Avg độ dài chunk",         int(sum(lengths)/len(lengths)) if lengths else 0),
        ("Chunk ≤ 512 ký tự",        sum(1 for l in lengths if l <= 512)),
        ("Chunk > 512 ký tự",        sum(1 for l in lengths if l > 512)),
    ]
    for i, (label, val) in enumerate(stats, 3):
        ws3[f"A{i}"] = label
        ws3[f"B{i}"] = val
        if label and not label.startswith(" ") and label:
            ws3[f"A{i}"].font = Font(bold=True)
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 15

    wb.save(output_path)
    n_dieu = sum(1 for c in chunks if c.get("chunk_sub_index", 0) == 0)
    log.info(f"\n✅ Excel chunks: {output_path}")
    log.info(f"   {total:,} chunks từ {n_dieu:,} điều")


# ── Xuất JSONL ────────────────────────────────────────────────────────────────

def export_jsonl(chunks: list, output_path: str):
    """Xuất JSONL chuẩn cho Qdrant / Chroma / Weaviate."""
    with open(output_path, "w", encoding="utf-8") as f:
        for chunk in chunks:
            obj = {
                "id":   chunk.get("chunk_id", ""),
                "text": chunk.get("context_text", ""),
                "payload": {
                    "source_file":       chunk.get("source_file", ""),
                    "ten_van_ban":       chunk.get("ten_van_ban", ""),
                    "so_hieu":           chunk.get("so_hieu", ""),
                    "loai_van_ban":      chunk.get("loai_van_ban", ""),
                    "trang_thai":        chunk.get("trang_thai", ""),
                    "ngay_hieu_luc":     chunk.get("ngay_hieu_luc", ""),
                    "ngay_het_hieu_luc": chunk.get("ngay_het_hieu_luc", ""),
                    "nhom":              chunk.get("nhom", ""),
                    "chuong_so":         chunk.get("chuong_so", ""),
                    "chuong_ten":        chunk.get("chuong_ten", ""),
                    "muc_so":            chunk.get("muc_so", ""),
                    "muc_ten":           chunk.get("muc_ten", ""),
                    "dieu_so":           chunk.get("dieu_so", ""),
                    "dieu_ten":          chunk.get("dieu_ten", ""),
                    "noi_dung_chunk":    chunk.get("noi_dung_chunk", ""),
                    "chunk_tier":        chunk.get("chunk_tier", 1),
                    "chunk_sub_index":   chunk.get("chunk_sub_index", 0),
                    "chunk_total_sub":   chunk.get("chunk_total_sub", 1),
                    "is_repealed":       chunk.get("is_repealed", False),
                    "is_truncated":      chunk.get("is_truncated", False),
                }
            }
            f.write(json.dumps(obj, ensure_ascii=False) + "\n")
    log.info(f"✅ JSONL: {output_path} ({len(chunks):,} chunks)")


# ── Summary ───────────────────────────────────────────────────────────────────

def print_summary(chunks: list):
    total = len(chunks)
    by_tier = {}
    for c in chunks:
        t = int(c.get("chunk_tier", 1))
        by_tier[t] = by_tier.get(t, 0) + 1
    repealed = sum(1 for c in chunks if c.get("is_repealed"))
    lengths  = [int(c.get("do_dai_chunk", 0)) for c in chunks if c.get("do_dai_chunk")]
    over_512 = sum(1 for l in lengths if l > 512)
    short    = sum(1 for l in lengths if l < 20)

    print("\n" + "="*60)
    print("  KẾT QUẢ CHUNKING")
    print("="*60)
    print(f"  Tổng chunks:               {total:>6,}")
    print(f"  Tier 1 - Nguyên (≤450):    {by_tier.get(1,0):>6,}")
    print(f"  Tier 2 - Split khoản:      {by_tier.get(2,0):>6,}")
    print(f"  Tier 3 - Split điểm:       {by_tier.get(3,0):>6,}")
    print(f"  Tier 4 - Recursive:        {by_tier.get(4,0):>6,}")
    print(f"  Tier 0 - Bãi bỏ:           {by_tier.get(0,0):>6,}")
    print(f"  Điều bị bãi bỏ:            {repealed:>6,}")
    if lengths:
        print(f"  Max độ dài chunk:          {max(lengths):>6,} ký tự")
        print(f"  Avg độ dài chunk:          {int(sum(lengths)/len(lengths)):>6,} ký tự")
        print(f"  Chunk > 512 ký tự còn lại:{over_512:>6,}  ← lý tưởng = 0")
        print(f"  Chunk < 20 ký tự (rác):   {short:>6,}  ← lý tưởng = 0")
    print("="*60 + "\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Smart chunker cho dữ liệu luật CNTT VN",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ví dụ:
  python smart_chunker.py --input data/law_data_output.xlsx --output data/law_chunks
  python smart_chunker.py --input data/law_data_output.xlsx --output data/law_chunks --format jsonl
        """
    )
    parser.add_argument("--input",       "-i", required=True, help="File Excel từ law_crawler.py")
    parser.add_argument("--output",      "-o", default="law_chunks", help="Đường dẫn output (không cần extension)")
    parser.add_argument("--format",      "-f", choices=["excel","jsonl","both"], default="both")
    parser.add_argument("--chunk_size",  type=int, default=CHUNK_SIZE,  help=f"Chunk size tối đa (mặc định {CHUNK_SIZE})")
    parser.add_argument("--overlap",     type=int, default=CHUNK_OVERLAP, help=f"Overlap (mặc định {CHUNK_OVERLAP})")
    parser.add_argument("--docx_folder", default=None, help="Folder DOCX gốc để patch điều bị truncate (tùy chọn)")
    args = parser.parse_args()

    # ── CẬP NHẬT CẤU HÌNH TOÀN CỤC ─────────────────────────────────────
    import smart_chunker as _m
    _m.CHUNK_SIZE    = args.chunk_size
    _m.CHUNK_OVERLAP = args.overlap

    # ── TỰ ĐỘNG DETECT FOLDER RAW (không cần gõ --docx_folder) ────────
    if not args.docx_folder:
        default_raw = Path("data/raw")
        if default_raw.exists():
            args.docx_folder = str(default_raw)
            log.info(f"🔧 Tự động detect --docx_folder = {default_raw} (để patch 2 điều truncate)")
        else:
            log.warning("⚠️ Không tìm thấy thư mục data/raw → 2 điều truncate có thể vẫn còn!")

    input_path = Path(args.input)
    if not input_path.exists():
        log.error(f"Không tìm thấy: {input_path}")
        sys.exit(1)

    log.info(f"📖 Đọc Excel: {input_path}")
    records = read_excel(str(input_path))

    # Patch điều bị truncate (nếu có folder raw)
    if args.docx_folder or any(r.get("is_truncated_excel") for r in records):
        records = patch_truncated_records(records, args.docx_folder)

    log.info(f"✂  Đang chunking {len(records):,} điều "
             f"(chunk_size={CHUNK_SIZE}, overlap={CHUNK_OVERLAP}, langchain={LANGCHAIN_OK})...")

    all_chunks = []
    for record in records:
        all_chunks.extend(chunk_record(record))

    # ── LOGGING MỚI: SỐ CHUNK THEO TỪNG VĂN BẢN (top 10) ───────────────
    from collections import defaultdict
    chunk_per_doc = defaultdict(int)
    for chunk in all_chunks:
        sf = chunk.get("source_file", "unknown")
        chunk_per_doc[sf] += 1

    print("\n" + "─" * 70)
    print("  SỐ CHUNK THEO TỪNG VĂN BẢN (top 10 nhiều chunk nhất)")
    print("─" * 70)
    for sf, cnt in sorted(chunk_per_doc.items(), key=lambda x: -x[1])[:10]:
        print(f"  {sf:<58} → {cnt:>6,} chunks")
    print(f"  Tổng cộng {len(chunk_per_doc)} văn bản")
    print("─" * 70)

    print_summary(all_chunks)

    # Tạo output path
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    stem = out.stem if out.suffix else out.name

    if args.format in ("excel", "both"):
        export_excel_chunks(all_chunks, str(out.parent / f"{stem}.xlsx"))
    if args.format in ("jsonl", "both"):
        export_jsonl(all_chunks, str(out.parent / f"{stem}.jsonl"))


if __name__ == "__main__":
    main()