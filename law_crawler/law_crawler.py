"""
law_crawler.py
==============
Crawl dữ liệu từ folder chứa file DOCX/PDF luật và nghị định,
parse ra cấu trúc: tên văn bản → chương → mục → điều → nội dung,
sau đó xuất ra file Excel với schema đầy đủ phục vụ pipeline RAG.

Stack: python-docx + openpyxl + re

Cách dùng:
    python law_crawler.py --input ./data/raw --output ./data/law_data_output.xlsx

Schema output:
    id_row | source_file | ten_van_ban | so_hieu | so_vbhn | loai_van_ban
    co_quan_ban_hanh | ngay_ban_hanh | ngay_hieu_luc | ngay_het_hieu_luc
    trang_thai | sua_doi_boi | ban_su_dung | nhom | ghi_chu
    chuong_so | chuong_ten | muc_so | muc_ten
    dieu_so | dieu_ten | noi_dung_dieu | do_dai_ky_tu | chunk_id
"""

import os
import re
import sys
import logging
import argparse
import hashlib
from pathlib import Path
from datetime import datetime
from typing import Optional

try:
    from docx import Document as DocxDocument
except ImportError:
    print("Thiếu thư viện: pip install python-docx")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Thiếu thư viện: pip install openpyxl")
    sys.exit(1)

# Import metadata config (cùng thư mục)
sys.path.insert(0, os.path.dirname(__file__))
from metadata_config import DOCUMENT_METADATA, FILENAME_KEYWORDS

# ── Logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger(__name__)

# ── Regex patterns ───────────────────────────────────────────────────────────
# Chương: "Chương I", "Chương 1", "CHƯƠNG I - ..."
RE_CHUONG = re.compile(
    r"^(chương\s+([IVXLCDM\d]+)[\s\-–—:.]*(.*)?)$",
    re.IGNORECASE
)
# Mục: "Mục 1", "MỤC 1 - ...", "Mục I"
RE_MUC = re.compile(
    r"^(mục\s+([IVXLCDM\d]+)[\s\-–—:.]*(.*)?)$",
    re.IGNORECASE
)
# Điều: "Điều 1.", "Điều 10.", "ĐIỀU 5."
RE_DIEU = re.compile(
    r"^(điều\s+(\d+)[\.:]?\s*(.*))$",
    re.IGNORECASE
)
# Khoản: "1.", "2.", v.v. - dùng để bỏ qua dòng chỉ là số
RE_KHOAN = re.compile(r"^\d+\.\s+")

# ── Pattern nhận diện phần KHÔNG phải nội dung điều ─────────────────────────
# Footnote separator: _____ (≥5 dấu _) → phần chú thích cuối VBHN
RE_FOOTNOTE_SEP  = re.compile(r"^_{5,}")
# Footnote item: [1], [2]... → chú thích được đánh số
RE_FOOTNOTE_ITEM = re.compile(r"^\[\d+\]")
# Citation dạng: "Điều X và Điều Y của Luật ABC" → trích dẫn điều từ luật khác
RE_CITATION_DIEU = re.compile(r"^điều\s+\d+\s+và\s+điều\s+\d+\s+của", re.IGNORECASE)
# Dòng ký xác thực cuối VBHN: "VĂN PHÒNG QUỐC HỘI", "XÁC THỰC VĂN BẢN"...
RE_VBHN_CLOSE = re.compile(r"^(văn phòng quốc hội|văn phòng chính phủ|xác thực văn bản|số:\s*\d|hà nội,|tp\.\s*hồ|chủ nhiệm|tổng thư ký|số:\s*\d+\/vbhn)", re.IGNORECASE)


# ── Helper functions ─────────────────────────────────────────────────────────

def normalize_text(text: str) -> str:
    """Chuẩn hoá khoảng trắng, xuống dòng."""
    text = text.strip()
    text = re.sub(r"\s+", " ", text)
    return text


def roman_to_int(s: str) -> int:
    """Chuyển số La Mã sang số nguyên để sort."""
    roman = {"I":1,"V":5,"X":10,"L":50,"C":100,"D":500,"M":1000}
    result, prev = 0, 0
    for ch in reversed(s.upper()):
        val = roman.get(ch, 0)
        result += val if val >= prev else -val
        prev = val
    return result


def extract_order_num(s: str) -> int:
    """Trích số thứ tự từ chuỗi (hỗ trợ La Mã và số Ả Rập)."""
    if s.isdigit():
        return int(s)
    try:
        return roman_to_int(s)
    except Exception:
        return 0


def make_chunk_id(source_file: str, dieu_so: str, index: int) -> str:
    """Tạo ID duy nhất cho từng điều (dùng cho vector DB)."""
    raw = f"{source_file}_{dieu_so}_{index}"
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def lookup_metadata(filename: str) -> dict:
    """
    Tìm metadata cho file dựa trên tên file.
    Ưu tiên khớp chính xác với key, sau đó dùng fuzzy keyword.
    """
    stem = Path(filename).stem.lower().strip()

    # Khớp trực tiếp key
    if stem in DOCUMENT_METADATA:
        return DOCUMENT_METADATA[stem]

    # Fuzzy: tìm keyword trong tên file
    for keyword, meta_key in FILENAME_KEYWORDS.items():
        if keyword in stem:
            return DOCUMENT_METADATA.get(meta_key, {})

    # Không tìm thấy -> trả về dict rỗng, log cảnh báo
    log.warning(f"  ⚠ Không tìm được metadata cho: {filename}")
    return {}


# ── Parser chính ─────────────────────────────────────────────────────────────

def parse_docx(filepath: str) -> list[dict]:
    """
    Parse file DOCX thành list các record theo cấp:
    chương → mục → điều → nội dung

    Returns:
        List[dict] - mỗi dict là một điều luật
    """
    try:
        doc = DocxDocument(filepath)
    except Exception as e:
        log.error(f"  ✗ Không đọc được file {filepath}: {e}")
        return []

    records = []
    filename = Path(filepath).name

    # State machine
    cur_chuong_so = ""
    cur_chuong_ten = ""
    cur_muc_so = ""
    cur_muc_ten = ""
    cur_dieu_so = ""
    cur_dieu_ten = ""
    cur_content_lines = []
    dieu_index = 0

    def flush_dieu():
        """Lưu điều hiện tại vào records."""
        nonlocal dieu_index
        if not cur_dieu_so:
            return
        # Lọc bỏ các dòng footnote bị lẫn vào nội dung điều
        # (trường hợp footnote không có dấu ___ phân cách)
        clean_lines = []
        for line in cur_content_lines:
            if RE_FOOTNOTE_SEP.match(line):
                break  # gặp ___ → dừng, bỏ phần sau
            if RE_FOOTNOTE_ITEM.match(line):
                continue  # bỏ [1] [2]... riêng lẻ
            clean_lines.append(line)
        content = " ".join(clean_lines).strip()
        content = re.sub(r"\s+", " ", content)
        if not content:
            return
        dieu_index += 1
        records.append({
            # ── Định danh file
            "source_file": filename,
            # ── Cấu trúc văn bản
            "chuong_so": cur_chuong_so,
            "chuong_ten": cur_chuong_ten,
            "muc_so": cur_muc_so,
            "muc_ten": cur_muc_ten,
            "dieu_so": cur_dieu_so,
            "dieu_ten": cur_dieu_ten,
            "noi_dung_dieu": content,
            # ── Kỹ thuật
            "do_dai_ky_tu": len(content),
            "chunk_id": make_chunk_id(filename, cur_dieu_so, dieu_index),
        })

    # Flag: đã vào vùng footnote/phụ lục cuối VBHN → bỏ qua hoàn toàn
    in_footnote_zone = False

    for para in doc.paragraphs:
        text = normalize_text(para.text)
        if not text:
            continue

        # ── Nhận diện và bỏ qua vùng footnote/phụ lục cuối VBHN ─────────────
        # Vùng này xuất hiện sau nội dung điều cuối cùng, bắt đầu bằng:
        #   (a) Đường kẻ ngang: ___________
        #   (b) Dòng xác thực: "VĂN PHÒNG QUỐC HỘI", "Hà Nội, ngày..."
        if RE_FOOTNOTE_SEP.match(text) or RE_VBHN_CLOSE.match(text):
            in_footnote_zone = True
            flush_dieu()  # lưu điều đang xử lý trước khi vào footnote
            cur_dieu_so = ""
            cur_content_lines = []
            continue

        if in_footnote_zone:
            continue  # bỏ qua toàn bộ nội dung trong vùng footnote

        # ── Bỏ qua dòng là footnote item đứng riêng lẻ [1] [2] ─────────────
        if RE_FOOTNOTE_ITEM.match(text):
            continue

        # ── Bỏ qua dòng là citation: "Điều X và Điều Y của Luật ABC" ────────
        # Đây là trích dẫn điều từ luật khác trong VBHN, không phải điều thật
        if RE_CITATION_DIEU.match(text):
            continue

        m_chuong = RE_CHUONG.match(text)
        m_muc = RE_MUC.match(text)
        m_dieu = RE_DIEU.match(text)

        if m_chuong:
            flush_dieu()
            cur_dieu_so = ""
            cur_dieu_ten = ""
            cur_content_lines = []
            cur_muc_so = ""
            cur_muc_ten = ""

            so_raw = m_chuong.group(2)
            ten = normalize_text(m_chuong.group(3)) if m_chuong.group(3) else ""
            cur_chuong_so = str(extract_order_num(so_raw)) if so_raw else so_raw
            cur_chuong_ten = ten

        elif m_muc:
            flush_dieu()
            cur_dieu_so = ""
            cur_dieu_ten = ""
            cur_content_lines = []

            so_raw = m_muc.group(2)
            ten = normalize_text(m_muc.group(3)) if m_muc.group(3) else ""
            cur_muc_so = str(extract_order_num(so_raw)) if so_raw else so_raw
            cur_muc_ten = ten

        elif m_dieu:
            flush_dieu()
            cur_dieu_so = m_dieu.group(2)
            cur_dieu_ten = normalize_text(m_dieu.group(3))
            # Nội dung bắt đầu từ dòng tiêu đề điều
            cur_content_lines = [text]

        else:
            # Nội dung thuộc điều đang xử lý
            if cur_dieu_so:
                cur_content_lines.append(text)

    # Flush điều cuối cùng
    flush_dieu()

    log.info(f"  ✓ {filename}: {len(records)} điều")
    return records


# ── Xuất Excel ───────────────────────────────────────────────────────────────

# Màu sắc header
COLOR_META   = "2E4057"   # Xanh đậm – metadata văn bản
COLOR_STRUCT = "048A81"   # Xanh lá – cấu trúc (chương/mục/điều)
COLOR_DATA   = "54C6EB"   # Xanh nhạt – nội dung kỹ thuật
COLOR_WHITE  = "FFFFFF"

# Định nghĩa các cột theo thứ tự
COLUMNS = [
    # (header, key_in_record, width, color_group)
    ("STT",              "stt",              6,  "meta"),
    ("Source File",      "source_file",      28, "meta"),
    ("Tên văn bản",      "ten_van_ban",      38, "meta"),
    ("Số hiệu",          "so_hieu",          18, "meta"),
    ("Số VBHN",          "so_vbhn",          18, "meta"),
    ("Loại văn bản",     "loai_van_ban",     16, "meta"),
    ("Cơ quan ban hành", "co_quan_ban_hanh", 22, "meta"),
    ("Ngày ban hành",    "ngay_ban_hanh",    16, "meta"),
    ("Ngày hiệu lực",    "ngay_hieu_luc",    16, "meta"),
    ("Ngày hết hiệu lực","ngay_het_hieu_luc",18, "meta"),
    ("Trạng thái",       "trang_thai",       18, "meta"),
    ("Sửa đổi bởi",      "sua_doi_boi",      35, "meta"),
    ("Bản sử dụng",      "ban_su_dung",      35, "meta"),
    ("Nhóm",             "nhom",             28, "meta"),
    ("Ghi chú",          "ghi_chu",          35, "meta"),
    ("Chương số",        "chuong_so",         12, "struct"),
    ("Chương tên",       "chuong_ten",        35, "struct"),
    ("Mục số",           "muc_so",            10, "struct"),
    ("Mục tên",          "muc_ten",           30, "struct"),
    ("Điều số",          "dieu_so",           10, "struct"),
    ("Điều tên",         "dieu_ten",          40, "struct"),
    ("Nội dung điều",    "noi_dung_dieu",     80, "data"),
    ("Độ dài (ký tự)",   "do_dai_ky_tu",      16, "data"),
    ("Chunk ID",         "chunk_id",          16, "data"),
]

COLOR_MAP = {
    "meta":   COLOR_META,
    "struct": COLOR_STRUCT,
    "data":   COLOR_DATA,
}


def make_header_fill(color_hex: str) -> PatternFill:
    return PatternFill("solid", fgColor=color_hex)


def make_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_excel(all_records: list[dict], output_path: str):
    """Xuất toàn bộ records ra file Excel đẹp, có header màu sắc."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Dữ liệu chính ───────────────────────────────────────────
    ws = wb.active
    ws.title = "Dữ liệu luật"
    ws.freeze_panes = "A2"  # Freeze header row

    border = make_border()

    # Header row
    for col_idx, (header, key, width, group) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        color = COLOR_MAP.get(group, COLOR_META)
        cell.font = Font(bold=True, color=COLOR_WHITE, size=11)
        cell.fill = make_header_fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, record in enumerate(all_records, start=2):
        # Màu xen kẽ nhẹ
        bg = "F7F9FC" if row_idx % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=bg)

        for col_idx, (header, key, width, group) in enumerate(COLUMNS, start=1):
            if key == "stt":
                value = row_idx - 1
            else:
                value = record.get(key, "")
                if value is None:
                    value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(key == "noi_dung_dieu")
            )

            # Highlight trạng thái
            if key == "trang_thai":
                if value == "con_hieu_luc":
                    cell.font = Font(color="1A7A44", bold=True)
                elif value == "het_hieu_luc":
                    cell.font = Font(color="C0392B", bold=True)
                elif value == "chua_hieu_luc":
                    cell.font = Font(color="D68910", bold=True)

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Sheet 2: Danh mục văn bản ────────────────────────────────────────
    ws2 = wb.create_sheet("Danh mục văn bản")
    ws2.freeze_panes = "A2"

    dmeta_cols = [
        ("STT", 6), ("Nhóm", 30), ("Tên văn bản", 40),
        ("Số hiệu", 18), ("Ngày hiệu lực", 16), ("Ngày hết hiệu lực", 18),
        ("Trạng thái", 18), ("Tổng số điều", 14), ("Ghi chú", 40)
    ]
    for ci, (h, w) in enumerate(dmeta_cols, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color=COLOR_WHITE, size=11)
        cell.fill = make_header_fill(COLOR_META)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 30

    # Đếm số điều theo văn bản
    count_by_file = {}
    for r in all_records:
        count_by_file[r["source_file"]] = count_by_file.get(r["source_file"], 0) + 1

    # Nhóm các metadata duy nhất theo source_file
    seen_files = {}
    for r in all_records:
        sf = r["source_file"]
        if sf not in seen_files:
            seen_files[sf] = r

    for ri, (sf, r) in enumerate(seen_files.items(), start=2):
        bg = "F7F9FC" if ri % 2 == 0 else "FFFFFF"
        fill2 = PatternFill("solid", fgColor=bg)
        vals = [
            ri - 1,
            r.get("nhom", ""),
            r.get("ten_van_ban", ""),
            r.get("so_hieu", ""),
            r.get("ngay_hieu_luc", ""),
            r.get("ngay_het_hieu_luc", "") or "—",
            r.get("trang_thai", ""),
            count_by_file.get(sf, 0),
            r.get("ghi_chu", "") or "",
        ]
        for ci2, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci2, value=v)
            c.fill = fill2
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(dmeta_cols))}1"

    # ── Sheet 3: Thống kê ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Thống kê")
    ws3["A1"] = "BÁO CÁO THỐNG KÊ DỮ LIỆU LUẬT CNTT VN"
    ws3["A1"].font = Font(bold=True, size=14, color=COLOR_META)
    ws3["A2"] = f"Ngày xử lý: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws3["A2"].font = Font(italic=True, color="888888")

    stats = [
        ("", ""),
        ("Tổng số điều đã crawl", len(all_records)),
        ("Tổng số văn bản", len(seen_files)),
        ("Số luật (Nhóm 1)", sum(1 for r in seen_files.values() if "Nhóm 1" in r.get("nhom", ""))),
        ("Số nghị định (Nhóm 2)", sum(1 for r in seen_files.values() if "Nhóm 2" in r.get("nhom", ""))),
        ("", ""),
        ("Theo trạng thái hiệu lực", ""),
        ("  Còn hiệu lực",     sum(1 for r in all_records if r.get("trang_thai") == "con_hieu_luc")),
        ("  Chưa có hiệu lực", sum(1 for r in all_records if r.get("trang_thai") == "chua_hieu_luc")),
        ("  Hết hiệu lực",     sum(1 for r in all_records if r.get("trang_thai") == "het_hieu_luc")),
        ("", ""),
        ("Top văn bản nhiều điều nhất", ""),
    ]
    # Top 5 văn bản nhiều điều
    top5 = sorted(count_by_file.items(), key=lambda x: x[1], reverse=True)[:5]
    for fname, cnt in top5:
        ten = seen_files.get(fname, {}).get("ten_van_ban", fname)
        stats.append((f"  {ten}", cnt))

    for ri, (label, value) in enumerate(stats, start=4):
        ws3[f"A{ri}"] = label
        ws3[f"B{ri}"] = value
        if label and not label.startswith(" ") and value != "":
            ws3[f"A{ri}"].font = Font(bold=True)

    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 20

    # Lưu file
    wb.save(output_path)
    log.info(f"\n✅ Đã xuất: {output_path}")
    log.info(f"   - Sheet 'Dữ liệu luật':      {len(all_records):,} dòng")
    log.info(f"   - Sheet 'Danh mục văn bản':  {len(seen_files)} văn bản")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Crawl DOCX luật/NĐ → Excel (DE tool)"
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Đường dẫn folder chứa file DOCX (hoặc 1 file DOCX)"
    )
    parser.add_argument(
        "--output", "-o",
        default="law_data_output.xlsx",
        help="Đường dẫn file Excel xuất ra (mặc định: law_data_output.xlsx)"
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Hiển thị chi tiết khi xử lý"
    )
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    input_path = Path(args.input)
    if not input_path.exists():
        log.error(f"Không tìm thấy: {input_path}")
        sys.exit(1)

    # Thu thập danh sách file DOCX
    if input_path.is_dir():
        docx_files = sorted(input_path.glob("**/*.docx"))
    elif input_path.suffix.lower() == ".docx":
        docx_files = [input_path]
    else:
        log.error("Chỉ hỗ trợ file .docx")
        sys.exit(1)

    if not docx_files:
        log.error(f"Không tìm thấy file .docx trong: {input_path}")
        sys.exit(1)

    log.info(f"📂 Tìm thấy {len(docx_files)} file DOCX")
    log.info("=" * 60)

    all_records = []
    for fp in docx_files:
        log.info(f"📄 Đang xử lý: {fp.name}")

        # Lấy metadata từ config
        meta = lookup_metadata(fp.name)

        # Parse điều luật
        raw_records = parse_docx(str(fp))

        # Gắn metadata vào từng record
        for rec in raw_records:
            rec.update(meta)
            # Đảm bảo source_file không bị ghi đè
            rec["source_file"] = fp.name

        all_records.extend(raw_records)

    log.info("=" * 60)
    log.info(f"📊 Tổng cộng: {len(all_records):,} điều từ {len(docx_files)} văn bản")
    log.info(f"💾 Đang xuất Excel: {args.output}")

    export_excel(all_records, args.output)


if __name__ == "__main__":
    main()